"""
Management command para importar datos de Ceibo desde Excel.

Mapeo Excel → DB:
  - Implementos: ID_Imp, Descripción Implemento, Accesorios, Nivel Rod
  - Familias: ID_FAM, Codigo, Familia, Implemento(nombre), Orden, Y/O, SI/NO
  - Productos: ID_Articulo, Cod Comercio, Plano, Cod Factura, Descripción,
               Implemento(nombre), Familia(nombre), Orden, _, Precio
  - Propiedades: ID_Prop, Descripción, Unidad
  - Prod_Prop: ID_ASIG, ID_Produ, ID_Imp, ID_Prop, Prioridad, _, _, _,
               Tipo(Exacto/Minimo/Maximo), Cantidad(valor), _, Neto(valor_neto)
  - Compat_Vetado: _, ID_Produ_PAD, ID_Produ_HIJ, _, _, _, _, Propiedad(tipo)
  - Prearmados (header row=1): ID_Prearmado, _, _, _, Prearmado(nombre),
                                Implemento(nombre), Precio Referencia
  - Estructuras: _, ID_Prearmado, ID_Producto, _, _, _, _, Cantidad
  - Precio_Productos: col K(10)=ID_Producto, col M(12)=Precio lista 81
"""

import os
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.catalogo.models import (
    Compatibilidad,
    Familia,
    Implemento,
    Producto,
    ProductoPropiedad,
    Propiedad,
)
from apps.clientes.models import FormaPago, TipoCliente
from apps.precios.models import (
    EstructuraPrearmado,
    ListaPrecio,
    Prearmado,
    PrecioProducto,
)
from apps.tenants.models import Tenant

# Agregación por propiedad según SPEC.md
AGREGACION_MAP = {
    'Longitud': 'SUM',
    'Peso': 'SUM',
    'Capacidad': 'SUM',
    'Altura': 'MAX',
    'Potencia': 'MAX',
    'LongitudLat': 'MAX',
    'Elásticos': 'MAX',
    'Ejes': 'MAX',
    'Llantas': 'MAX',
    'Centro': 'MAX',
}

DEFAULT_FILE = os.path.join('scripts', 'data', 'DBA_Cotizador_Ceibo.xlsx')


def to_decimal(val, default=Decimal('0')):
    if val is None:
        return default
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return default


def to_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return default


def is_empty_row(row, check_col=0):
    val = row[check_col] if len(row) > check_col else None
    return val is None or str(val).strip() in ('', '-')


class Command(BaseCommand):
    help = 'Importar datos de Ceibo desde el archivo Excel DBA_Cotizador_Ceibo.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default=DEFAULT_FILE,
            help='Ruta al archivo Excel',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        filepath = options['file']
        if not os.path.exists(filepath):
            raise CommandError(f'Archivo no encontrado: {filepath}')

        self.stdout.write(f'Cargando {filepath}...')
        wb = openpyxl.load_workbook(filepath, data_only=True)

        tenant = self._create_tenant()
        imp_map = self._import_implementos(wb['Implementos'], tenant)
        fam_map = self._import_familias(wb['Familias'], tenant, imp_map)
        prod_map = self._import_productos(wb['Productos'], tenant, imp_map, fam_map)
        prop_map = self._import_propiedades(wb['Propiedades'], tenant)
        self._import_producto_propiedades(wb['Prod_Prop'], prod_map, prop_map)
        self._import_compatibilidades(wb['Compat_Vetado'], tenant, prod_map)
        prearmado_map = self._import_prearmados(wb['Prearmados'], tenant, imp_map)
        self._import_estructuras(wb['Estructuras'], prearmado_map, prod_map)
        self._import_precios(wb['Precio_Productos'], wb['Productos'], tenant, prod_map)
        self._create_tipos_cliente_y_formas_pago(tenant)

        wb.close()
        self.stdout.write(self.style.SUCCESS('Importación completada exitosamente.'))

    def _create_tenant(self):
        tenant, created = Tenant.objects.get_or_create(
            slug='ceibo',
            defaults={
                'nombre': 'Metalúrgica Ceibo',
                'bonif_max_porcentaje': Decimal('30.00'),
                'moneda': 'ARS',
            },
        )
        action = 'Creado' if created else 'Ya existía'
        self.stdout.write(f'  Tenant Ceibo: {action}')
        return tenant

    def _import_implementos(self, ws, tenant):
        self.stdout.write('Importando Implementos...')
        imp_map = {}  # excel_id → Implemento
        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_empty_row(row, 1):
                break
            excel_id = to_int(row[0])
            nombre = str(row[1]).strip()
            accesorios = str(row[2]).strip() if row[2] and str(row[2]).strip() != '-' else None
            nivel_rodado = to_int(row[3]) if row[3] and str(row[3]).strip() not in ('', '-') else None

            imp, _ = Implemento.objects.get_or_create(
                tenant=tenant,
                nombre=nombre,
                defaults={
                    'accesorios_tipo': accesorios,
                    'nivel_rodado': nivel_rodado,
                },
            )
            imp_map[excel_id] = imp
            imp_map[nombre] = imp  # also by name for lookups
        self.stdout.write(f'  {len([k for k in imp_map if isinstance(k, int)])} implementos importados')
        return imp_map

    def _import_familias(self, ws, tenant, imp_map):
        self.stdout.write('Importando Familias...')
        fam_map = {}  # excel_id → Familia, also name → Familia
        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_empty_row(row, 2):
                break
            excel_id = to_int(row[0])
            nombre = str(row[2]).strip()
            imp_nombre = str(row[3]).strip()
            orden = to_int(row[4])
            tipo_sel = str(row[5]).strip() if row[5] else 'O'
            obligatoria = str(row[6]).strip() if row[6] else 'SI'

            implemento = imp_map.get(imp_nombre)
            if not implemento:
                self.stderr.write(f'  WARN: Implemento "{imp_nombre}" no encontrado para familia "{nombre}"')
                continue

            fam, _ = Familia.objects.get_or_create(
                tenant=tenant,
                implemento=implemento,
                nombre=nombre,
                defaults={
                    'orden': orden,
                    'tipo_seleccion': tipo_sel,
                    'obligatoria': obligatoria,
                },
            )
            fam_map[excel_id] = fam
            fam_map[nombre] = fam
        self.stdout.write(f'  {len([k for k in fam_map if isinstance(k, int)])} familias importadas')
        return fam_map

    def _import_productos(self, ws, tenant, imp_map, fam_map):
        self.stdout.write('Importando Productos...')
        prod_map = {}  # excel_id → Producto
        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_empty_row(row, 0):
                break
            excel_id = to_int(row[0])
            cod_comercio = str(row[1]).strip() if row[1] and str(row[1]).strip() != '-' else None
            plano = str(row[2]).strip() if row[2] and str(row[2]).strip() != '-' else None
            cod_factura = str(row[3]).strip() if row[3] and str(row[3]).strip() != '-' else None
            nombre = str(row[4]).strip()
            imp_nombre = str(row[5]).strip()
            fam_nombre = str(row[6]).strip()
            orden = to_int(row[7])

            implemento = imp_map.get(imp_nombre)
            familia = fam_map.get(fam_nombre)

            if not implemento:
                self.stderr.write(f'  WARN: Implemento "{imp_nombre}" no encontrado para producto "{nombre}"')
                continue
            if not familia:
                self.stderr.write(f'  WARN: Familia "{fam_nombre}" no encontrada para producto "{nombre}"')
                continue

            prod, _ = Producto.objects.get_or_create(
                tenant=tenant,
                nombre=nombre,
                familia=familia,
                defaults={
                    'implemento': implemento,
                    'cod_comercio': cod_comercio,
                    'plano': plano,
                    'cod_factura': cod_factura,
                    'orden': orden,
                    'iva_porcentaje': Decimal('21.00'),
                },
            )
            prod_map[excel_id] = prod
        self.stdout.write(f'  {len(prod_map)} productos importados')
        return prod_map

    def _import_propiedades(self, ws, tenant):
        self.stdout.write('Importando Propiedades...')
        prop_map = {}  # excel_id → Propiedad
        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_empty_row(row, 1):
                break
            excel_id = to_int(row[0])
            nombre = str(row[1]).strip()
            unidad = str(row[2]).strip()
            agregacion = AGREGACION_MAP.get(nombre, 'SUM')

            prop, _ = Propiedad.objects.get_or_create(
                tenant=tenant,
                nombre=nombre,
                defaults={
                    'unidad': unidad,
                    'agregacion': agregacion,
                },
            )
            prop_map[excel_id] = prop
        self.stdout.write(f'  {len(prop_map)} propiedades importadas')
        return prop_map

    def _import_producto_propiedades(self, ws, prod_map, prop_map):
        self.stdout.write('Importando ProductoPropiedades...')
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_empty_row(row, 0):
                break
            prod_excel_id = to_int(row[1])
            prop_excel_id = to_int(row[3])
            prioridad = to_int(row[4])
            tipo = str(row[8]).strip() if row[8] else 'Exacto'
            valor = to_decimal(row[9])
            valor_neto_raw = row[11]
            valor_neto = to_decimal(valor_neto_raw) if valor_neto_raw and str(valor_neto_raw).strip() not in ('', '-') else None

            producto = prod_map.get(prod_excel_id)
            propiedad = prop_map.get(prop_excel_id)

            if not producto:
                continue
            if not propiedad:
                continue

            ProductoPropiedad.objects.get_or_create(
                producto=producto,
                propiedad=propiedad,
                tipo=tipo,
                defaults={
                    'valor': valor,
                    'valor_neto': valor_neto,
                    'prioridad': prioridad,
                },
            )
            count += 1
        self.stdout.write(f'  {count} producto-propiedades importadas')

    def _import_compatibilidades(self, ws, tenant, prod_map):
        self.stdout.write('Importando Compatibilidades...')
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_empty_row(row, 0):
                break
            padre_id = to_int(row[1])
            hijo_id = to_int(row[2])
            tipo_raw = str(row[7]).strip() if len(row) > 7 and row[7] else 'Vetado'

            padre = prod_map.get(padre_id)
            hijo = prod_map.get(hijo_id)

            if not padre or not hijo:
                continue

            Compatibilidad.objects.get_or_create(
                tenant=tenant,
                producto_padre=padre,
                producto_hijo=hijo,
                defaults={'tipo': tipo_raw},
            )
            count += 1
        self.stdout.write(f'  {count} compatibilidades importadas')

    def _import_prearmados(self, ws, tenant, imp_map):
        self.stdout.write('Importando Prearmados...')
        prearmado_map = {}  # excel_id → Prearmado
        # Header is row 1 (index 1), data starts at row 2 (index 2)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if is_empty_row(row, 0):
                break
            excel_id = to_int(row[0])
            nombre = str(row[4]).strip()
            imp_nombre = str(row[5]).strip()
            precio_ref = to_decimal(row[6]) if row[6] and str(row[6]).strip() not in ('', '-') else None

            implemento = imp_map.get(imp_nombre)
            if not implemento:
                self.stderr.write(f'  WARN: Implemento "{imp_nombre}" no encontrado para prearmado "{nombre}"')
                continue

            pre, _ = Prearmado.objects.get_or_create(
                tenant=tenant,
                nombre=nombre,
                defaults={
                    'implemento': implemento,
                    'precio_referencia': precio_ref,
                },
            )
            prearmado_map[excel_id] = pre
        self.stdout.write(f'  {len(prearmado_map)} prearmados importados')
        return prearmado_map

    def _import_estructuras(self, ws, prearmado_map, prod_map):
        self.stdout.write('Importando Estructuras de Prearmados...')
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_empty_row(row, 0):
                break
            prearmado_excel_id = to_int(row[1])
            producto_excel_id = to_int(row[2])
            cantidad = to_int(row[7], default=1)

            prearmado = prearmado_map.get(prearmado_excel_id)
            producto = prod_map.get(producto_excel_id)

            if not prearmado or not producto:
                continue

            EstructuraPrearmado.objects.get_or_create(
                prearmado=prearmado,
                producto=producto,
                defaults={'cantidad': cantidad},
            )
            count += 1
        self.stdout.write(f'  {count} estructuras importadas')

    def _import_precios(self, ws_precios, ws_productos, tenant, prod_map):
        self.stdout.write('Importando Lista de Precios #81...')

        lista, _ = ListaPrecio.objects.get_or_create(
            tenant=tenant,
            numero=81,
            defaults={
                'nombre': 'Lista Ceibo Octubre 2025',
                'estado': 'vigente',
            },
        )

        # Usar precios de la hoja Productos (col 9 = Precio)
        count = 0
        for row in ws_productos.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            excel_id = to_int(row[0])
            precio_raw = row[9]

            producto = prod_map.get(excel_id)
            if not producto:
                continue

            precio = to_decimal(precio_raw)

            PrecioProducto.objects.get_or_create(
                lista=lista,
                producto=producto,
                defaults={'precio': precio},
            )
            count += 1

        self.stdout.write(f'  Lista #81 creada con {count} precios')

    def _create_tipos_cliente_y_formas_pago(self, tenant):
        self.stdout.write('Creando Tipos de Cliente y Formas de Pago...')

        tipos = [
            ('Concesionario', Decimal('15.00')),
            ('Vendedor', Decimal('10.00')),
            ('Cliente Final', Decimal('0.00')),
        ]
        for nombre, bonif in tipos:
            TipoCliente.objects.get_or_create(
                tenant=tenant,
                nombre=nombre,
                defaults={'bonificacion_default': bonif},
            )

        formas = [
            ('Contado', Decimal('15.00')),
            ('Cheque 30 días', Decimal('10.00')),
            ('Cheque 60 días', Decimal('5.00')),
            ('Financiado', Decimal('0.00')),
        ]
        for nombre, bonif in formas:
            FormaPago.objects.get_or_create(
                tenant=tenant,
                nombre=nombre,
                defaults={'bonificacion_porcentaje': bonif},
            )

        self.stdout.write(f'  {len(tipos)} tipos de cliente, {len(formas)} formas de pago')
